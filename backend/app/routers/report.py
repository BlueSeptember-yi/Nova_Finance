"""报表管理路由"""

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.account import Account
from app.models.journal import JournalEntry, LedgerLine
from app.utils.auth import get_current_user
from app.utils.helpers import success_response

router = APIRouter(prefix="/reports", tags=["报表管理"])


@router.get("/income-statement", response_model=dict)
def generate_income_statement(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """生成利润表"""

    # 1. 获取营业收入（主营业务收入 6001）
    revenue_account = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code == "6001",
        )
        .first()
    )

    revenue = Decimal(0)
    if revenue_account:
        revenue_result = (
            db.query(func.sum(LedgerLine.credit) - func.sum(LedgerLine.debit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == revenue_account.account_id,
                Account.company_id == current_user.company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        revenue = revenue_result if revenue_result else Decimal(0)

    # 2. 获取营业成本（主营业务成本 6401）
    cost_account = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code == "6401",
        )
        .first()
    )

    cost = Decimal(0)
    if cost_account:
        cost_result = (
            db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == cost_account.account_id,
                Account.company_id == current_user.company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        cost = cost_result if cost_result else Decimal(0)

    # 3. 获取期间费用（销售费用6601 + 管理费用6602 + 财务费用6603）
    expense_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code.in_(["6601", "6602", "6603"]),
        )
        .all()
    )

    expenses = Decimal(0)
    if expense_accounts:
        for expense_account in expense_accounts:
            expense_result = (
                db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
                .join(Account)
                .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
                .filter(
                    Account.account_id == expense_account.account_id,
                    Account.company_id == current_user.company_id,
                    JournalEntry.date >= start_date,
                    JournalEntry.date <= end_date,
                    JournalEntry.posted == True,
                )
                .scalar()
            )
            if expense_result:
                expenses += expense_result

    # 4. 获取税金及附加（税金及附加 6403）
    tax_account = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code == "6403",  # 税金及附加费用账户
        )
        .first()
    )

    tax = Decimal(0)
    if tax_account:
        tax_result = (
            db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == tax_account.account_id,
                Account.company_id == current_user.company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        tax = tax_result if tax_result else Decimal(0)

    # 5. 计算营业利润和净利润
    operating_profit = revenue - cost - expenses
    net_profit = operating_profit - tax

    result = {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "revenue": float(revenue),
        "cost": float(cost),
        "expenses": float(expenses),
        "operating_profit": float(operating_profit),
        "tax": float(tax),
        "net_profit": float(net_profit),
    }

    return success_response(data=result, message="利润表生成成功")


def calculate_account_balance(
    account: Account, db: Session, as_of_date: Optional[date] = None
) -> Decimal:
    """计算账户余额（考虑余额方向，包含所有子科目的余额）"""

    # 获取所有子科目（递归）
    def get_all_child_accounts(parent_account_id: str) -> list:
        """递归获取所有子科目ID"""
        children = (
            db.query(Account).filter(Account.parent_id == parent_account_id).all()
        )
        result = [parent_account_id]
        for child in children:
            result.extend(get_all_child_accounts(child.account_id))
        return result

    # 获取当前科目及其所有子科目的ID列表
    account_ids = get_all_child_accounts(account.account_id)

    if account.normal_balance == "Debit":
        # 借方余额账户：余额 = 借方 - 贷方
        if as_of_date:
            # 计算到指定日期的余额（包含所有子科目）
            debit_sum = db.query(func.sum(LedgerLine.debit)).join(
                JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
            ).filter(
                LedgerLine.account_id.in_(account_ids),
                JournalEntry.posted == True,
                JournalEntry.date <= as_of_date,
            ).scalar() or Decimal(0)
            credit_sum = db.query(func.sum(LedgerLine.credit)).join(
                JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
            ).filter(
                LedgerLine.account_id.in_(account_ids),
                JournalEntry.posted == True,
                JournalEntry.date <= as_of_date,
            ).scalar() or Decimal(0)
            return debit_sum - credit_sum
        else:
            # 使用缓存的余额（包含所有子科目）
            total_debit = Decimal(0)
            total_credit = Decimal(0)
            for acc_id in account_ids:
                acc = db.query(Account).filter(Account.account_id == acc_id).first()
                if acc:
                    total_debit += acc.balance_debit or Decimal(0)
                    total_credit += acc.balance_credit or Decimal(0)
            return total_debit - total_credit
    else:
        # 贷方余额账户：余额 = 贷方 - 借方
        if as_of_date:
            debit_sum = db.query(func.sum(LedgerLine.debit)).join(
                JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
            ).filter(
                LedgerLine.account_id.in_(account_ids),
                JournalEntry.posted == True,
                JournalEntry.date <= as_of_date,
            ).scalar() or Decimal(0)
            credit_sum = db.query(func.sum(LedgerLine.credit)).join(
                JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
            ).filter(
                LedgerLine.account_id.in_(account_ids),
                JournalEntry.posted == True,
                JournalEntry.date <= as_of_date,
            ).scalar() or Decimal(0)
            return credit_sum - debit_sum
        else:
            # 使用缓存的余额（包含所有子科目）
            total_debit = Decimal(0)
            total_credit = Decimal(0)
            for acc_id in account_ids:
                acc = db.query(Account).filter(Account.account_id == acc_id).first()
                if acc:
                    total_debit += acc.balance_debit or Decimal(0)
                    total_credit += acc.balance_credit or Decimal(0)
            return total_credit - total_debit


def build_account_tree(accounts: list, parent_id: str = None) -> list:
    """构建科目树形结构"""
    tree = []
    for account in accounts:
        if account.parent_id == parent_id:
            node = {
                "account_id": account.account_id,
                "code": account.code,
                "name": account.name,
                "balance": None,  # 将在外部计算
                "children": build_account_tree(accounts, account.account_id),
            }
            tree.append(node)
    return tree


@router.get("/balance-sheet", response_model=dict)
def generate_balance_sheet(
    as_of_date: date = Query(..., description="报表日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """生成资产负债表"""

    # 1. 计算资产（Asset类型账户）
    # 只查询父科目（parent_id 为 NULL）
    asset_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.type == "Asset",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    # 获取所有资产类科目（包括子科目）用于构建树形结构
    all_asset_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.type == "Asset",
        )
        .all()
    )

    assets_list = []
    total_assets = Decimal(0)

    # 流动资产
    current_assets = Decimal(0)
    # 非流动资产
    non_current_assets = Decimal(0)

    for account in asset_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            # 构建该科目的树形结构
            account_tree = {
                "account_id": account.account_id,
                "code": account.code,
                "name": account.name,
                "balance": float(balance),
                "children": [],
            }

            # 添加子科目（仅用于显示，不重复计算余额）
            children = (
                db.query(Account)
                .filter(
                    Account.company_id == current_user.company_id,
                    Account.parent_id == account.account_id,
                )
                .all()
            )
            for child in children:
                child_balance = calculate_account_balance(child, db, as_of_date)
                if child_balance != 0:
                    account_tree["children"].append(
                        {
                            "account_id": child.account_id,
                            "code": child.code,
                            "name": child.name,
                            "balance": float(child_balance),
                            "children": [],
                        }
                    )

            assets_list.append(account_tree)
            total_assets += balance

            # 根据科目编码判断是流动资产还是非流动资产
            code_int = (
                int(account.code.split(".")[0])
                if "." in account.code
                else int(account.code)
            )
            if code_int < 1600:  # 1600以下通常是流动资产
                current_assets += balance
            else:
                non_current_assets += balance

    # 2. 计算负债（Liability类型账户）
    # 只查询父科目（parent_id 为 NULL）
    liability_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.type == "Liability",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    liabilities_list = []
    total_liabilities = Decimal(0)

    # 流动负债
    current_liabilities = Decimal(0)
    # 非流动负债
    non_current_liabilities = Decimal(0)

    for account in liability_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            # 构建该科目的树形结构
            account_tree = {
                "account_id": account.account_id,
                "code": account.code,
                "name": account.name,
                "balance": float(balance),
                "children": [],
            }

            # 添加子科目（仅用于显示，不重复计算余额）
            children = (
                db.query(Account)
                .filter(
                    Account.company_id == current_user.company_id,
                    Account.parent_id == account.account_id,
                )
                .all()
            )
            for child in children:
                child_balance = calculate_account_balance(child, db, as_of_date)
                if child_balance != 0:
                    account_tree["children"].append(
                        {
                            "account_id": child.account_id,
                            "code": child.code,
                            "name": child.name,
                            "balance": float(child_balance),
                            "children": [],
                        }
                    )

            liabilities_list.append(account_tree)
            total_liabilities += balance

            # 根据科目编码判断是流动负债还是非流动负债
            code_int = (
                int(account.code.split(".")[0])
                if "." in account.code
                else int(account.code)
            )
            if code_int < 2500:  # 2500以下通常是流动负债
                current_liabilities += balance
            else:
                non_current_liabilities += balance

    # 3. 计算所有者权益（Equity类型账户）
    # 只查询父科目（parent_id 为 NULL）
    equity_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.type == "Equity",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    equity_list = []
    total_equity = Decimal(0)

    for account in equity_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            # 构建该科目的树形结构
            account_tree = {
                "account_id": account.account_id,
                "code": account.code,
                "name": account.name,
                "balance": float(balance),
                "children": [],
            }

            # 添加子科目（仅用于显示，不重复计算余额）
            children = (
                db.query(Account)
                .filter(
                    Account.company_id == current_user.company_id,
                    Account.parent_id == account.account_id,
                )
                .all()
            )
            for child in children:
                child_balance = calculate_account_balance(child, db, as_of_date)
                if child_balance != 0:
                    account_tree["children"].append(
                        {
                            "account_id": child.account_id,
                            "code": child.code,
                            "name": child.name,
                            "balance": float(child_balance),
                            "children": [],
                        }
                    )

            equity_list.append(account_tree)
            total_equity += balance

    # 4. 计算本年利润（从利润表计算）
    # 获取本年利润账户（4103）
    profit_account = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code == "4103",
        )
        .first()
    )

    current_year_profit = Decimal(0)
    if profit_account:
        current_year_profit = calculate_account_balance(profit_account, db, as_of_date)

    # 如果本年利润账户没有余额，尝试从损益类账户计算
    if current_year_profit == 0:
        # 计算所有收入类账户的余额
        revenue_accounts = (
            db.query(Account)
            .filter(
                Account.company_id == current_user.company_id,
                Account.type == "Revenue",
            )
            .all()
        )

        total_revenue = Decimal(0)
        for account in revenue_accounts:
            # 收入类账户是贷方余额
            if as_of_date:
                credit_sum = db.query(
                    func.sum(LedgerLine.credit) - func.sum(LedgerLine.debit)
                ).join(
                    JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
                ).filter(
                    LedgerLine.account_id == account.account_id,
                    JournalEntry.posted == True,
                    JournalEntry.date <= as_of_date,
                ).scalar() or Decimal(0)
                total_revenue += credit_sum

        # 计算所有费用类账户的余额
        expense_accounts = (
            db.query(Account)
            .filter(
                Account.company_id == current_user.company_id,
                Account.type == "Expense",
            )
            .all()
        )

        total_expense = Decimal(0)
        for account in expense_accounts:
            # 费用类账户是借方余额
            if as_of_date:
                debit_sum = db.query(
                    func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit)
                ).join(
                    JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
                ).filter(
                    LedgerLine.account_id == account.account_id,
                    JournalEntry.posted == True,
                    JournalEntry.date <= as_of_date,
                ).scalar() or Decimal(0)
                total_expense += debit_sum

        current_year_profit = total_revenue - total_expense

    total_equity += current_year_profit

    # 5. 验证平衡：资产 = 负债 + 所有者权益
    total_liabilities_and_equity = total_liabilities + total_equity
    balance_check = abs(total_assets - total_liabilities_and_equity)
    is_balanced = balance_check < Decimal("0.01")  # 允许0.01的误差

    result = {
        "as_of_date": as_of_date.isoformat(),
        "assets": {
            "current_assets": float(current_assets),
            "non_current_assets": float(non_current_assets),
            "total": float(total_assets),
            "details": assets_list,  # 使用树形结构
        },
        "liabilities": {
            "current_liabilities": float(current_liabilities),
            "non_current_liabilities": float(non_current_liabilities),
            "total": float(total_liabilities),
            "details": liabilities_list,  # 使用树形结构
        },
        "equity": {
            "total": float(total_equity),
            "current_year_profit": float(current_year_profit),
            "details": equity_list,  # 使用树形结构
        },
        "total_liabilities_and_equity": float(total_liabilities_and_equity),
        "balance_check": float(balance_check),
        "is_balanced": is_balanced,
    }

    return success_response(data=result, message="资产负债表生成成功")


@router.get("/cash-flow", response_model=dict)
def generate_cash_flow(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """生成现金流量表"""

    # 1. 经营活动产生的现金流量
    # 销售商品、提供劳务收到的现金（主要来自应收账款减少和主营业务收入）
    operating_cash_in = Decimal(0)

    # 购买商品、接受劳务支付的现金（主要来自应付账款减少和主营业务成本）
    operating_cash_out = Decimal(0)

    # 获取现金类账户（1001库存现金, 1002银行存款）
    cash_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == current_user.company_id,
            Account.code.in_(["1001", "1002"]),
        )
        .all()
    )

    # 计算经营活动现金流（通过分析现金账户的变动）
    for cash_account in cash_accounts:
        # 获取该期间内所有涉及现金账户的分录
        cash_lines = (
            db.query(LedgerLine)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                LedgerLine.account_id == cash_account.account_id,
                JournalEntry.posted == True,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
            )
            .all()
        )

        for line in cash_lines:
            # 现金增加（借方）
            if line.debit > 0:
                # 判断是否来自经营活动
                journal = line.journal_entry
                if journal.source_type in ["SO", "RECEIPT", "MANUAL"]:
                    # 检查对应的贷方科目
                    other_lines = [
                        l
                        for l in journal.lines
                        if l.account_id != cash_account.account_id
                    ]
                    for other_line in other_lines:
                        other_account = (
                            db.query(Account)
                            .filter(Account.account_id == other_line.account_id)
                            .first()
                        )
                        if other_account:
                            if (
                                other_account.type == "Revenue"
                                or other_account.code == "1122"
                            ):  # 应收账款
                                operating_cash_in += line.debit

            # 现金减少（贷方）
            if line.credit > 0:
                # 判断是否来自经营活动
                journal = line.journal_entry
                if journal.source_type in ["PO", "PAYMENT", "MANUAL"]:
                    # 检查对应的借方科目
                    other_lines = [
                        l
                        for l in journal.lines
                        if l.account_id != cash_account.account_id
                    ]
                    for other_line in other_lines:
                        other_account = (
                            db.query(Account)
                            .filter(Account.account_id == other_line.account_id)
                            .first()
                        )
                        if other_account:
                            if (
                                other_account.type == "Expense"
                                or other_account.code == "2201"
                            ):  # 应付账款
                                operating_cash_out += line.credit

    # 2. 投资活动产生的现金流量（简化处理）
    investing_cash_in = Decimal(0)
    investing_cash_out = Decimal(0)

    # 3. 筹资活动产生的现金流量（简化处理）
    financing_cash_in = Decimal(0)
    financing_cash_out = Decimal(0)

    # 计算净现金流
    operating_net = operating_cash_in - operating_cash_out
    investing_net = investing_cash_in - investing_cash_out
    financing_net = financing_cash_in - financing_cash_out
    net_cash_flow = operating_net + investing_net + financing_net

    # 计算期初和期末现金余额
    beginning_cash = Decimal(0)
    ending_cash = Decimal(0)

    for cash_account in cash_accounts:
        # 期初余额（start_date之前）
        beginning_balance = calculate_account_balance(cash_account, db, start_date)
        beginning_cash += beginning_balance

        # 期末余额（end_date）
        ending_balance = calculate_account_balance(cash_account, db, end_date)
        ending_cash += ending_balance

    result = {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "operating_activities": {
            "cash_in": float(operating_cash_in),
            "cash_out": float(operating_cash_out),
            "net": float(operating_net),
        },
        "investing_activities": {
            "cash_in": float(investing_cash_in),
            "cash_out": float(investing_cash_out),
            "net": float(investing_net),
        },
        "financing_activities": {
            "cash_in": float(financing_cash_in),
            "cash_out": float(financing_cash_out),
            "net": float(financing_net),
        },
        "net_cash_flow": float(net_cash_flow),
        "beginning_cash": float(beginning_cash),
        "ending_cash": float(ending_cash),
    }

    return success_response(data=result, message="现金流量表生成成功")


def _get_income_statement_data(
    start_date: date,
    end_date: date,
    company_id: str,
    db: Session,
) -> dict:
    """获取利润表数据（内部函数，用于导出）"""
    # 1. 获取营业收入（主营业务收入 6001）
    revenue_account = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code == "6001",
        )
        .first()
    )

    revenue = Decimal(0)
    if revenue_account:
        revenue_result = (
            db.query(func.sum(LedgerLine.credit) - func.sum(LedgerLine.debit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == revenue_account.account_id,
                Account.company_id == company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        revenue = revenue_result if revenue_result else Decimal(0)

    # 2. 获取营业成本（主营业务成本 6401）
    cost_account = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code == "6401",
        )
        .first()
    )

    cost = Decimal(0)
    if cost_account:
        cost_result = (
            db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == cost_account.account_id,
                Account.company_id == company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        cost = cost_result if cost_result else Decimal(0)

    # 3. 获取期间费用
    expense_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code.in_(["6601", "6602", "6603"]),
        )
        .all()
    )

    expenses = Decimal(0)
    for expense_account in expense_accounts:
        expense_result = (
            db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == expense_account.account_id,
                Account.company_id == company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        if expense_result:
            expenses += expense_result

    # 4. 获取税金及附加（税金及附加 6403）
    tax_account = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code == "6403",  # 税金及附加费用账户
        )
        .first()
    )

    tax = Decimal(0)
    if tax_account:
        tax_result = (
            db.query(func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit))
            .join(Account)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                Account.account_id == tax_account.account_id,
                Account.company_id == company_id,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
                JournalEntry.posted == True,
            )
            .scalar()
        )
        tax = tax_result if tax_result else Decimal(0)

    # 5. 计算营业利润和净利润
    operating_profit = revenue - cost - expenses
    net_profit = operating_profit - tax

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "revenue": float(revenue),
        "cost": float(cost),
        "expenses": float(expenses),
        "operating_profit": float(operating_profit),
        "tax": float(tax),
        "net_profit": float(net_profit),
    }


@router.get("/export/income-statement")
def export_income_statement(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出利润表（Excel）"""
    # 生成利润表数据
    data = _get_income_statement_data(start_date, end_date, current_user.company_id, db)

    # 创建Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 准备数据
        df_data = [
            {"项目": "一、营业收入", "金额": data["revenue"]},
            {"项目": "减：营业成本", "金额": data["cost"]},
            {"项目": "减：期间费用", "金额": data["expenses"]},
            {"项目": "二、营业利润", "金额": data["operating_profit"]},
            {"项目": "减：税金及附加", "金额": data["tax"]},
            {"项目": "三、净利润", "金额": data["net_profit"]},
        ]

        df = pd.DataFrame(df_data)
        df.to_excel(writer, sheet_name="利润表", index=False, startrow=3)

        worksheet = writer.sheets["利润表"]

        # 设置列宽
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 20

        # 导入样式
        from openpyxl.styles import Alignment, Font, PatternFill, numbers

        # 添加标题
        worksheet.merge_cells("A1:B1")
        worksheet["A1"] = "利润表"
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 添加日期范围
        worksheet.merge_cells("A2:B2")
        worksheet["A2"] = f"报表期间：{start_date} 至 {end_date}"
        worksheet["A2"].font = Font(size=12)
        worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # 设置表头样式
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[4]:  # 第4行是表头（因为startrow=3）
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置数据格式
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=2, max_col=2
        ):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER_00
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # 设置项目列对齐
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=1, max_col=1
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 加粗关键行
        bold_font = Font(bold=True)
        worksheet["A5"].font = bold_font  # 营业收入
        worksheet["A8"].font = bold_font  # 营业利润
        worksheet["A10"].font = bold_font  # 净利润
        worksheet["B5"].font = bold_font
        worksheet["B8"].font = bold_font
        worksheet["B10"].font = bold_font

    output.seek(0)
    filename = f"利润表_{start_date}_{end_date}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


def _get_balance_sheet_data(
    as_of_date: date,
    company_id: str,
    db: Session,
) -> dict:
    """获取资产负债表数据（内部函数，用于导出）"""
    # 1. 计算资产（只计算父科目）
    asset_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.type == "Asset",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    assets_list = []
    total_assets = Decimal(0)
    current_assets = Decimal(0)
    non_current_assets = Decimal(0)

    for account in asset_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            assets_list.append(
                {
                    "code": account.code,
                    "name": account.name,
                    "balance": float(balance),
                }
            )
            total_assets += balance
            code_int = (
                int(account.code.split(".")[0])
                if "." in account.code
                else int(account.code)
            )
            if code_int < 1600:
                current_assets += balance
            else:
                non_current_assets += balance

    # 2. 计算负债（只计算父科目）
    liability_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.type == "Liability",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    liabilities_list = []
    total_liabilities = Decimal(0)
    current_liabilities = Decimal(0)
    non_current_liabilities = Decimal(0)

    for account in liability_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            liabilities_list.append(
                {
                    "code": account.code,
                    "name": account.name,
                    "balance": float(balance),
                }
            )
            total_liabilities += balance
            code_int = (
                int(account.code.split(".")[0])
                if "." in account.code
                else int(account.code)
            )
            if code_int < 2500:
                current_liabilities += balance
            else:
                non_current_liabilities += balance

    # 3. 计算所有者权益（只计算父科目）
    equity_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.type == "Equity",
            Account.parent_id.is_(None),  # 只计算父科目
        )
        .all()
    )

    equity_list = []
    total_equity = Decimal(0)

    for account in equity_accounts:
        balance = calculate_account_balance(account, db, as_of_date)
        if balance != 0:
            equity_list.append(
                {
                    "code": account.code,
                    "name": account.name,
                    "balance": float(balance),
                }
            )
            total_equity += balance

    # 4. 计算本年利润
    profit_account = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code == "4103",
        )
        .first()
    )

    current_year_profit = Decimal(0)
    if profit_account:
        current_year_profit = calculate_account_balance(profit_account, db, as_of_date)

    if current_year_profit == 0:
        revenue_accounts = (
            db.query(Account)
            .filter(
                Account.company_id == company_id,
                Account.type == "Revenue",
            )
            .all()
        )

        total_revenue = Decimal(0)
        for account in revenue_accounts:
            if as_of_date:
                credit_sum = db.query(
                    func.sum(LedgerLine.credit) - func.sum(LedgerLine.debit)
                ).join(
                    JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
                ).filter(
                    LedgerLine.account_id == account.account_id,
                    JournalEntry.posted == True,
                    JournalEntry.date <= as_of_date,
                ).scalar() or Decimal(0)
                total_revenue += credit_sum

        expense_accounts = (
            db.query(Account)
            .filter(
                Account.company_id == company_id,
                Account.type == "Expense",
            )
            .all()
        )

        total_expense = Decimal(0)
        for account in expense_accounts:
            if as_of_date:
                debit_sum = db.query(
                    func.sum(LedgerLine.debit) - func.sum(LedgerLine.credit)
                ).join(
                    JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id
                ).filter(
                    LedgerLine.account_id == account.account_id,
                    JournalEntry.posted == True,
                    JournalEntry.date <= as_of_date,
                ).scalar() or Decimal(0)
                total_expense += debit_sum

        current_year_profit = total_revenue - total_expense

    total_equity += current_year_profit
    total_liabilities_and_equity = total_liabilities + total_equity

    return {
        "as_of_date": as_of_date.isoformat(),
        "assets": {
            "current_assets": float(current_assets),
            "non_current_assets": float(non_current_assets),
            "total": float(total_assets),
            "details": assets_list,  # 只包含父科目
        },
        "liabilities": {
            "current_liabilities": float(current_liabilities),
            "non_current_liabilities": float(non_current_liabilities),
            "total": float(total_liabilities),
            "details": liabilities_list,  # 只包含父科目
        },
        "equity": {
            "total": float(total_equity),
            "current_year_profit": float(current_year_profit),
            "details": equity_list,  # 只包含父科目
        },
        "total_liabilities_and_equity": float(total_liabilities_and_equity),
    }


@router.get("/export/balance-sheet")
def export_balance_sheet(
    as_of_date: date = Query(..., description="报表日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出资产负债表（Excel）"""
    # 生成资产负债表数据
    data = _get_balance_sheet_data(as_of_date, current_user.company_id, db)

    # 创建Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 资产部分
        assets_data = []
        assets_data.append(
            {"项目": "流动资产", "金额": data["assets"]["current_assets"]}
        )
        for detail in data["assets"]["details"]:
            if detail["balance"] != 0:
                assets_data.append(
                    {"项目": f"  {detail['name']}", "金额": detail["balance"]}
                )
        assets_data.append(
            {"项目": "非流动资产", "金额": data["assets"]["non_current_assets"]}
        )
        assets_data.append({"项目": "资产合计", "金额": data["assets"]["total"]})

        # 负债和所有者权益部分
        liabilities_data = []
        liabilities_data.append(
            {"项目": "流动负债", "金额": data["liabilities"]["current_liabilities"]}
        )
        for detail in data["liabilities"]["details"]:
            if detail["balance"] != 0:
                liabilities_data.append(
                    {"项目": f"  {detail['name']}", "金额": detail["balance"]}
                )
        liabilities_data.append(
            {
                "项目": "非流动负债",
                "金额": data["liabilities"]["non_current_liabilities"],
            }
        )
        liabilities_data.append(
            {"项目": "负债合计", "金额": data["liabilities"]["total"]}
        )

        equity_data = []
        for detail in data["equity"]["details"]:
            if detail["balance"] != 0:
                equity_data.append({"项目": detail["name"], "金额": detail["balance"]})
        equity_data.append(
            {"项目": "本年利润", "金额": data["equity"]["current_year_profit"]}
        )
        equity_data.append({"项目": "所有者权益合计", "金额": data["equity"]["total"]})
        equity_data.append(
            {
                "项目": "负债和所有者权益合计",
                "金额": data["total_liabilities_and_equity"],
            }
        )

        # 合并数据
        max_len = max(len(assets_data), len(liabilities_data) + len(equity_data))
        df_data = []
        for i in range(max_len):
            row = {}
            if i < len(assets_data):
                row["资产"] = assets_data[i]["项目"]
                row["资产金额"] = assets_data[i]["金额"]
            if i < len(liabilities_data):
                row["负债和所有者权益"] = liabilities_data[i]["项目"]
                row["负债和所有者权益金额"] = liabilities_data[i]["金额"]
            elif i - len(liabilities_data) < len(equity_data):
                idx = i - len(liabilities_data)
                row["负债和所有者权益"] = equity_data[idx]["项目"]
                row["负债和所有者权益金额"] = equity_data[idx]["金额"]
            df_data.append(row)

        df = pd.DataFrame(df_data)
        df.to_excel(writer, sheet_name="资产负债表", index=False, startrow=3)

        worksheet = writer.sheets["资产负债表"]

        # 设置列宽
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 20
        worksheet.column_dimensions["C"].width = 30
        worksheet.column_dimensions["D"].width = 20

        # 导入样式
        from openpyxl.styles import Alignment, Font, PatternFill, numbers

        # 添加标题
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "资产负债表"
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 添加日期
        worksheet.merge_cells("A2:D2")
        worksheet["A2"] = f"报表日期：{as_of_date}"
        worksheet["A2"].font = Font(size=12)
        worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # 设置表头样式
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[4]:  # 第4行是表头
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置数据格式
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=2, max_col=2
        ):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=4, max_col=4
        ):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # 设置项目列对齐
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=1, max_col=1
        ):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=3, max_col=3
        ):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 加粗合计行
        bold_font = Font(bold=True)
        for row_idx in range(5, worksheet.max_row + 1):
            if worksheet[f"A{row_idx}"].value and (
                "合计" in str(worksheet[f"A{row_idx}"].value)
                or "合计" in str(worksheet[f"C{row_idx}"].value)
            ):
                for col in ["A", "B", "C", "D"]:
                    cell = worksheet[f"{col}{row_idx}"]
                    if cell.value is not None:
                        cell.font = bold_font

    output.seek(0)
    filename = f"资产负债表_{as_of_date}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


def _get_cash_flow_data(
    start_date: date,
    end_date: date,
    company_id: str,
    db: Session,
) -> dict:
    """获取现金流量表数据（内部函数，用于导出）"""
    # 使用与generate_cash_flow相同的逻辑
    operating_cash_in = Decimal(0)
    operating_cash_out = Decimal(0)

    cash_accounts = (
        db.query(Account)
        .filter(
            Account.company_id == company_id,
            Account.code.in_(["1001", "1002"]),
        )
        .all()
    )

    for cash_account in cash_accounts:
        cash_lines = (
            db.query(LedgerLine)
            .join(JournalEntry, LedgerLine.journal_id == JournalEntry.journal_id)
            .filter(
                LedgerLine.account_id == cash_account.account_id,
                JournalEntry.posted == True,
                JournalEntry.date >= start_date,
                JournalEntry.date <= end_date,
            )
            .all()
        )

        for line in cash_lines:
            if line.debit > 0:
                journal = line.journal_entry
                if journal.source_type in ["SO", "RECEIPT", "MANUAL"]:
                    other_lines = [
                        l
                        for l in journal.lines
                        if l.account_id != cash_account.account_id
                    ]
                    for other_line in other_lines:
                        other_account = (
                            db.query(Account)
                            .filter(Account.account_id == other_line.account_id)
                            .first()
                        )
                        if other_account:
                            if (
                                other_account.type == "Revenue"
                                or other_account.code == "1122"
                            ):
                                operating_cash_in += line.debit

            if line.credit > 0:
                journal = line.journal_entry
                if journal.source_type in ["PO", "PAYMENT", "MANUAL"]:
                    other_lines = [
                        l
                        for l in journal.lines
                        if l.account_id != cash_account.account_id
                    ]
                    for other_line in other_lines:
                        other_account = (
                            db.query(Account)
                            .filter(Account.account_id == other_line.account_id)
                            .first()
                        )
                        if other_account:
                            if (
                                other_account.type == "Expense"
                                or other_account.code == "2201"
                            ):
                                operating_cash_out += line.credit

    investing_cash_in = Decimal(0)
    investing_cash_out = Decimal(0)
    financing_cash_in = Decimal(0)
    financing_cash_out = Decimal(0)

    operating_net = operating_cash_in - operating_cash_out
    investing_net = investing_cash_in - investing_cash_out
    financing_net = financing_cash_in - financing_cash_out
    net_cash_flow = operating_net + investing_net + financing_net

    beginning_cash = Decimal(0)
    ending_cash = Decimal(0)

    for cash_account in cash_accounts:
        beginning_balance = calculate_account_balance(cash_account, db, start_date)
        beginning_cash += beginning_balance
        ending_balance = calculate_account_balance(cash_account, db, end_date)
        ending_cash += ending_balance

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "operating_activities": {
            "cash_in": float(operating_cash_in),
            "cash_out": float(operating_cash_out),
            "net": float(operating_net),
        },
        "investing_activities": {
            "cash_in": float(investing_cash_in),
            "cash_out": float(investing_cash_out),
            "net": float(investing_net),
        },
        "financing_activities": {
            "cash_in": float(financing_cash_in),
            "cash_out": float(financing_cash_out),
            "net": float(financing_net),
        },
        "net_cash_flow": float(net_cash_flow),
        "beginning_cash": float(beginning_cash),
        "ending_cash": float(ending_cash),
    }


@router.get("/export/cash-flow")
def export_cash_flow(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出现金流量表（Excel）"""
    # 生成现金流量表数据
    data = _get_cash_flow_data(start_date, end_date, current_user.company_id, db)

    # 创建Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 准备数据
        df_data = [
            {"项目": "一、经营活动产生的现金流量", "金额": None},
            {"项目": "  现金流入", "金额": data["operating_activities"]["cash_in"]},
            {"项目": "  现金流出", "金额": data["operating_activities"]["cash_out"]},
            {
                "项目": "  经营活动产生的现金流量净额",
                "金额": data["operating_activities"]["net"],
            },
            {"项目": "", "金额": None},
            {"项目": "二、投资活动产生的现金流量", "金额": None},
            {"项目": "  现金流入", "金额": data["investing_activities"]["cash_in"]},
            {"项目": "  现金流出", "金额": data["investing_activities"]["cash_out"]},
            {
                "项目": "  投资活动产生的现金流量净额",
                "金额": data["investing_activities"]["net"],
            },
            {"项目": "", "金额": None},
            {"项目": "三、筹资活动产生的现金流量", "金额": None},
            {"项目": "  现金流入", "金额": data["financing_activities"]["cash_in"]},
            {"项目": "  现金流出", "金额": data["financing_activities"]["cash_out"]},
            {
                "项目": "  筹资活动产生的现金流量净额",
                "金额": data["financing_activities"]["net"],
            },
            {"项目": "", "金额": None},
            {"项目": "四、现金及现金等价物净增加额", "金额": data["net_cash_flow"]},
            {"项目": "  期初现金及现金等价物余额", "金额": data["beginning_cash"]},
            {"项目": "  期末现金及现金等价物余额", "金额": data["ending_cash"]},
        ]

        df = pd.DataFrame(df_data)
        df.to_excel(writer, sheet_name="现金流量表", index=False, startrow=3)

        worksheet = writer.sheets["现金流量表"]

        # 设置列宽
        worksheet.column_dimensions["A"].width = 35
        worksheet.column_dimensions["B"].width = 20

        # 导入样式
        from openpyxl.styles import Alignment, Font, PatternFill, numbers

        # 添加标题
        worksheet.merge_cells("A1:B1")
        worksheet["A1"] = "现金流量表"
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 添加日期范围
        worksheet.merge_cells("A2:B2")
        worksheet["A2"] = f"报表期间：{start_date} 至 {end_date}"
        worksheet["A2"].font = Font(size=12)
        worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # 设置表头样式
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[4]:  # 第4行是表头
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置数据格式
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=2, max_col=2
        ):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = numbers.FORMAT_NUMBER_00
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # 设置项目列对齐
        for row in worksheet.iter_rows(
            min_row=5, max_row=worksheet.max_row, min_col=1, max_col=1
        ):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 加粗关键行
        bold_font = Font(bold=True)
        key_rows = [5, 8, 11, 14, 20, 21, 22]  # 主要项目行
        for row_idx in key_rows:
            if row_idx <= worksheet.max_row:
                for col in ["A", "B"]:
                    cell = worksheet[f"{col}{row_idx}"]
                    if cell.value is not None:
                        cell.font = bold_font

    output.seek(0)
    filename = f"现金流量表_{start_date}_{end_date}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )
